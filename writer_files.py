import datetime
import os.path
from collections import defaultdict

import pandas as pd

from openpyxl import Workbook, load_workbook

def write_to_excel_row_station(year, month, region, data, list_path):
    header = list(data.values())[0]

    if len(header) == 0:
        return

    pd_data = pd.DataFrame(data, index=list_path)
    with (pd.ExcelWriter(f'row_station\\{year}_{month}_{region}.xlsx') as writer):
        pd_data.to_excel(writer, index=True)


def write_to_excel(data, table_const, dates):
    early_date, lates_date = map(lambda d: datetime.datetime.strptime(d, "%m.%Y"), dates[:2])
    file_table_stations = table_const.set_index('Horizon')

    data_from_write = defaultdict()
    list_regions = []
    list_dataframes = []
    if data.mean_result_data_begin:
        for region in data.mean_result_data_begin:
            data_from_write[f'T1 ({data.mean_early_date})'] = data.mean_result_data_begin[region]
            data_from_write[f'T2 ({data.mean_lates_date})'] = data.mean_result_data_end[region]
            data_from_write['ΔT'] = data.delta_t[region]
            data_from_write['Tср'] = data.mean_delta_t[region]
            data_from_write['V'] = file_table_stations[f'V, {region}']
            data_from_write['S'] = file_table_stations[f'S, {region}']
            data_from_write['Q'] = data.q_count_termal[region]
            data_from_write['Qsum'] = data.q_sum_down[region]
            data_from_write['Q/S'] = data.q_per_s[region]
            data_from_write[f'T1_gradients ({data.mean_early_date})'] = data.t_gradients_1[region]
            data_from_write[f'T2_gradients ({data.mean_lates_date})'] = data.t_gradients_2[region]
            data_from_write['T_gradients_ср'] = data.t_gradients_mean[region]
            data_from_write['Kz'] = data.k_z[region]
            dataframe = pd.DataFrame(data_from_write)
            list_regions.append(region)
            list_dataframes.append(dataframe)
        with pd.ExcelWriter(
                f'result\\{early_date.year}_{early_date.month}-{lates_date.year}_{lates_date.month}.xlsx') as writer:
            for region, dataframe in zip(list_regions, list_dataframes):
                dataframe.to_excel(writer, sheet_name = region, index = True)

def write_to_excel_all_kz(data, table_const, dates):
    early_date, lates_date = map(lambda d: datetime.datetime.strptime(d, "%m.%Y"), dates[:2])
    horizons = table_const['Horizon']

    path = 'result\\Data.xlsx'
    if os.path.exists(path):
        book = load_workbook(path)
    else:
        book = Workbook()

    for region, region_kzs in data.k_z.items():
        if region in book.worksheets:
            page = book.worksheets[region]
        else:
            # если листа не существовало, создаем и заполняем горизонтами
            page = book.create_sheet(region)
            # заголовок столбца
            row = 1
            page.cell(row,1).value = "Горизонт"
            row += 1
            # значения горизонтов
            for horizon in horizons:
                page.cell(row, 1).value = horizon
                row += 1

        # добавляем дополнительный столбец
        kz_column = page.max_column + 1
        # заголовок столбца
        page.cell(1, kz_column).value = f"Kz {early_date.year}_{early_date.month}-{lates_date.year}_{lates_date.month}"
        # заполняем столбец
        for row in range(2, page.max_row):
            # получаем значение горизонта в текущей строке
            curr_hrz = int(page.cell(row, 1).value)
            # проверяем, сесть ли данные для этого горизонта
            kz = region_kzs.get(curr_hrz)
            # если есть - записываем
            if kz:
                page.cell(row, kz_column).value = kz

    book.save(path)

def write_statistic_station(year, month, region, list_path, name):
    count = len(list_path)

    with open('statistics\\statistics_station.txt', 'a', encoding='utf-8') as file:
        if count:
            file.write(f'{year}, {month}, {region}, {count}, {name} \n')